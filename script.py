import pdfplumber
import openpyxl

# Define file paths (Users must replace these with their own file paths)
pdf_path = "path/to/input.pdf"  # Path to the source PDF
excel_path = "path/to/output.xlsx"  # Path to the destination Excel file

# Function to extract data from a specific page of the PDF
def extract_data_from_pdf(pdf_path, page_number):
    extracted_data = []
    
    with pdfplumber.open(pdf_path) as pdf:
        if page_number < len(pdf.pages):
            page = pdf.pages[page_number]
            table = page.extract_table()
            
            if table:
                extracted_data = table  # Store extracted table data
            else:
                print("No table found on the specified page.")
        else:
            print("Invalid page number.")
    
    return extracted_data

# Function to update the Excel file with extracted PDF data
def update_excel_file(excel_path, extracted_data):
    try:
        # Load the existing Excel file
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active  # Use the first sheet
        
        # Ensure extracted data is not empty before modifying the sheet
        if extracted_data:
            sheet.cell(row=11, column=3, value=extracted_data[0][0])  # Example: Update row 11, column 3
            sheet.cell(row=16, column=5, value=extracted_data[1][1])  # Example: Update row 16, column 5
        
            # Save the updated Excel file
            wb.save(excel_path)
            print("Excel file updated successfully.")
        else:
            print("No data extracted. Excel file remains unchanged.")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

# Extract data from the PDF
page_number = 2  # Update this based on where your table is located
extracted_data = extract_data_from_pdf(pdf_path, page_number)

# Update the Excel file with the extracted data
update_excel_file(excel_path, extracted_data)

print("Process completed.")
